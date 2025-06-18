import streamlit as st
import pandas as pd
from app.core.db_connector import get_db_engine
from sqlalchemy import text
import datetime
import io # For BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF # For PDF export

# --- Page Configuration ---
st.set_page_config(page_title="資產負債表 (Balance Sheet)", layout="wide")
st.title("資產負債表 (Balance Sheet)")

# --- Database Connection ---
@st.cache_resource
def init_db_connection():
    engine = get_db_engine()
    return engine

db_engine = init_db_connection()
if not db_engine:
    st.error("資料庫引擎未能初始化。請檢查您的資料庫配置。")
    st.stop()

# --- Helper function to fetch data ---
@st.cache_data(ttl=3600) # Cache company data longer
def fetch_company_data_bs(): # Renamed for Balance Sheet page
    if not db_engine:
        return pd.DataFrame(columns=['CO_NO', 'CO_NAME'])
    try:
        with db_engine.connect() as connection:
            df = pd.read_sql(text("SELECT CO_NO, CO_NAME FROM PCOMPANY ORDER BY CO_NO"), connection)
            return df
    except Exception as e:
        st.error(f"獲取公司列表時發生錯誤 (資產負債表): {e}")
        return pd.DataFrame(columns=['CO_NO', 'CO_NAME'])

@st.cache_data(ttl=300)
def execute_query(query, params=None):
    try:
        with db_engine.connect() as connection:
            df = pd.read_sql(text(query), connection, params=params)
            return df
    except Exception as e:
        st.error(f"資料查詢時發生錯誤: {e}")
        return pd.DataFrame()

# --- Balance Sheet Account Categories & Structure ---
BS_STRUCTURE = {
    "Assets": {"pattern_prefix": "1", "display_name": "資產 (Assets)"},
    "Liabilities": {"pattern_prefix": "2", "display_name": "負債 (Liabilities)"},
    "Equity": {"pattern_prefix": "3", "display_name": "權益 (Equity)"}
}

def get_account_balances(as_of_date_str, account_pattern_prefix, company_no): # Added company_no
    if not company_no:
        st.sidebar.error("資產負債表錯誤：未選擇公司進行查詢。")
        return pd.DataFrame(columns=['科目代號 (Account Code)', '科目名稱 (Account Name)', '金額 (Amount)'])

    query = """
    SELECT
        a.AT_NO, a.AT_NAME, a.AT_DCR,
        COALESCE(SUM(CASE
            WHEN d.SD_DOC = 'D' THEN d.SD_AMT
            WHEN d.SD_DOC = 'C' THEN -d.SD_AMT
            ELSE 0
        END), 0) AS NetDebitAmount
    FROM AACNT a
    LEFT JOIN ASPDT d ON a.AT_NO = d.SD_ATNO
    LEFT JOIN ASLIP h ON d.SD_NO = h.SP_NO AND d.SD_INDEX = h.SP_INDEX AND h.SP_CHECK = '1' AND h.SP_DATE <= :as_of_date
    WHERE a.AT_NO LIKE :pattern AND h.SP_CO_NO = :company_no -- Added company filter
    GROUP BY a.AT_NO, a.AT_NAME, a.AT_DCR
    HAVING COALESCE(SUM(CASE WHEN d.SD_DOC = 'D' THEN d.SD_AMT WHEN d.SD_DOC = 'C' THEN -d.SD_AMT ELSE 0 END), 0) != 0
    ORDER BY a.AT_NO;
    """
    params = {'as_of_date': as_of_date_str, 'pattern': f"{account_pattern_prefix}%", 'company_no': company_no}
    df = execute_query(query, params)
    if df.empty:
        return pd.DataFrame(columns=['科目代號 (Account Code)', '科目名稱 (Account Name)', '金額 (Amount)'])
    df['Balance'] = df.apply(lambda row: row['NetDebitAmount'] if row['AT_DCR'] == 'D' else -row['NetDebitAmount'], axis=1)
    return df[['AT_NO', 'AT_NAME', 'Balance']].rename(columns={
        'AT_NO': '科目代號 (Account Code)', 'AT_NAME': '科目名稱 (Account Name)', 'Balance': '金額 (Amount)'
    })

def generate_balance_sheet_df(as_of_date, company_no): # Added company_no
    as_of_date_str = as_of_date.strftime('%Y%m%d')
    bs_data_frames = []
    section_totals = {}

    for section_key, section_details in BS_STRUCTURE.items():
        bs_data_frames.append(pd.DataFrame([{'科目代號 (Account Code)': section_details["display_name"], '科目名稱 (Account Name)': '', '金額 (Amount)': None}]))
        accounts_df = get_account_balances(as_of_date_str, section_details["pattern_prefix"], company_no) # Pass company_no
        if not accounts_df.empty:
            bs_data_frames.append(accounts_df)
        section_total = accounts_df['金額 (Amount)'].sum() if not accounts_df.empty else 0
        section_totals[section_key] = section_total
        bs_data_frames.append(pd.DataFrame([{'科目代號 (Account Code)': f"{section_details['display_name']} 總計", '科目名稱 (Account Name)': '', '金額 (Amount)': section_total}]))
        bs_data_frames.append(pd.DataFrame([{'科目代號 (Account Code)': '', '科目名稱 (Account Name)': '', '金額 (Amount)': None}])) # Spacer

    final_bs_df = pd.concat(bs_data_frames, ignore_index=True).fillna('') # Fill NaN with empty string for display consistency

    total_liabilities = section_totals.get('Liabilities', 0)
    total_equity = section_totals.get('Equity', 0)
    total_liabilities_and_equity = total_liabilities + total_equity

    # Append Total L+E row to the DataFrame for consistent export
    final_bs_df = pd.concat([final_bs_df, pd.DataFrame([
        {'科目代號 (Account Code)': "負債及權益總計 (Total Liabilities and Equity)", '科目名稱 (Account Name)': '', '金額 (Amount)': total_liabilities_and_equity}
    ])], ignore_index=True)

    return final_bs_df, section_totals

# --- Export Functions ---
def bs_df_to_excel(df, params):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"BalanceSheet_{params['as_of_date'].strftime('%Y%m%d')}"

    ws.append([f"截至 {params['as_of_date'].strftime('%Y-%m-%d')} 資產負債表"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1] if df.shape[1] > 0 else 1)
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])

    header = [cell for cell in df.columns]
    ws.append(header)
    header_row_num = 3
    for cell in ws[header_row_num:header_row_num]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for r_idx, row_values in enumerate(dataframe_to_rows(df, index=False, header=False), header_row_num + 1):
        ws.append(row_values)
        # Style section headers and total rows
        # Assuming first column contains item names like "資產 (Assets)", "資產 (Assets) 總計"
        if isinstance(row_values[0], str) and ("總計" in row_values[0] or row_values[0] in [s["display_name"] for s in BS_STRUCTURE.values()] or "負債及權益總計" in row_values[0]):
            for cell in ws[r_idx:r_idx]:
                cell.font = Font(bold=True)

    # Number formatting and column widths
    amount_col_idx = df.columns.get_loc('金額 (Amount)') + 1 if '金額 (Amount)' in df.columns else -1
    if amount_col_idx != -1:
        amount_col_letter = get_column_letter(amount_col_idx)
        for r in range(header_row_num + 1, ws.max_row + 1):
            cell_value = ws[f'{amount_col_letter}{r}'].value
            if isinstance(cell_value, (int, float)):
                ws[f'{amount_col_letter}{r}'].number_format = '#,##0'
        ws.column_dimensions[amount_col_letter].width = 18

    ws.column_dimensions[get_column_letter(1)].width = 45 # Account Code/Name
    if df.shape[1] > 1:
         ws.column_dimensions[get_column_letter(2)].width = 30 # Account Name

    wb.save(output)
    output.seek(0)
    return output.getvalue()

def bs_df_to_pdf(df, params):
    pdf = FPDF()
    pdf.add_page()
    try:
        pdf.add_font('uming', '', '/usr/share/fonts/truetype/arphic/uming.ttc', uni=True)
        pdf.set_font('uming', '', 10)
    except RuntimeError:
        try:
            pdf.add_font('msjh', '', 'C:/Windows/Fonts/msjh.ttc', uni=True) # For Windows
            pdf.set_font('msjh', '', 10)
        except RuntimeError:
            st.warning("CJK font not found for PDF export. Chinese characters might not render correctly.")
            pdf.set_font('Arial', '', 10)

    pdf.cell(0, 10, f"截至 {params['as_of_date'].strftime('%Y-%m-%d')} 資產負債表", 0, 1, 'C')

    header = list(df.columns)
    col_widths = {'科目代號 (Account Code)': 100, '科目名稱 (Account Name)': 60, '金額 (Amount)': 30}
    line_height = pdf.font_size * 1.5

    for idx, h_text in enumerate(header):
        pdf.cell(col_widths.get(h_text, 40), line_height, h_text, border=1, ln=0 if idx < len(header)-1 else 1, align='C')

    for _, row_series in df.iterrows():
        for idx, (col_name, cell_value) in enumerate(row_series.items()):
            align = 'R' if col_name == '金額 (Amount)' and cell_value not in ["", None] and isinstance(cell_value, str) and cell_value.replace(',','').replace('-','').isdigit() else 'L'
            # cell_value can be None or empty string from fillna('')
            cell_text = str(cell_value) if cell_value is not None else ''
            pdf.cell(col_widths.get(col_name, 40), line_height, cell_text, border=1, ln=0 if idx < len(header)-1 else 1, align=align)

    return pdf.output(dest='S').encode('latin-1')

# --- Filter Section ---
st.sidebar.header("報表參數")

# Company Selector
companies_df_bs = fetch_company_data_bs()
company_options_bs = {row['CO_NO']: f"{row['CO_NO']} - {row['CO_NAME']}" for _, row in companies_df_bs.iterrows()} if not companies_df_bs.empty else {}
selected_company_no_bs = None
selected_company_name_bs = "無公司"

if not company_options_bs:
    st.sidebar.warning("未找到任何公司資料 (資產負債表)。")
else:
    default_company_no_bs = list(company_options_bs.keys())[0]
    selected_company_no_bs = st.sidebar.selectbox(
        "公司 (Company)", options=list(company_options_bs.keys()),
        format_func=lambda x: company_options_bs.get(x, "未知公司"),
        key="bs_company_no", index=0
    )
    selected_company_name_bs = company_options_bs.get(selected_company_no_bs, "未知公司")

default_bs_date = datetime.date.today()
bs_as_of_date = st.sidebar.date_input("選擇截止日期 (As of Date)", value=default_bs_date, key="bs_as_of_date")

# --- Session State ---
if 'balance_sheet_display_df' not in st.session_state:
    st.session_state.balance_sheet_display_df = pd.DataFrame()
if 'balance_sheet_raw_df' not in st.session_state:
    st.session_state.balance_sheet_raw_df = pd.DataFrame()
if 'balance_sheet_totals' not in st.session_state:
    st.session_state.balance_sheet_totals = {}
if 'balance_sheet_params' not in st.session_state:
    st.session_state.balance_sheet_params = None

# --- Generate Report Button ---
if st.sidebar.button("生成報表", type="primary", key="generate_bs_button"):
    if not selected_company_no_bs:
        st.error("請選擇一個公司。")
    elif not bs_as_of_date:
        st.warning("請選擇截止日期。")
    else:
        st.session_state.balance_sheet_params = {
            "as_of_date": bs_as_of_date,
            "company_no": selected_company_no_bs, # Store selected company
            "company_name": selected_company_name_bs
        }
        with st.spinner(f"正在為 {selected_company_name_bs} 生成截至 {bs_as_of_date.strftime('%Y-%m-%d')} 的資產負債表..."): # Updated spinner
            raw_df, totals = generate_balance_sheet_df(bs_as_of_date, selected_company_no_bs) # Pass company_no
            st.session_state.balance_sheet_raw_df = raw_df.copy()
            st.session_state.balance_sheet_totals = totals

            display_df = raw_df.copy()
            # Ensure '金額 (Amount)' column exists before trying to format it
            if '金額 (Amount)' in display_df.columns:
                display_df['金額 (Amount)'] = display_df['金額 (Amount)'].apply(
                    lambda x: f"{x:,.0f}" if pd.notnull(x) and isinstance(x, (int, float)) else (x if isinstance(x, str) else '')
                )
            st.session_state.balance_sheet_display_df = display_df

# --- Display Area for Balance Sheet ---
if not st.session_state.balance_sheet_display_df.empty and st.session_state.balance_sheet_params:
    params = st.session_state.balance_sheet_params
    totals = st.session_state.balance_sheet_totals

    st.subheader(f"{params.get('company_name','')} - 截至 {params['as_of_date'].strftime('%Y-%m-%d')} 資產負債表") # Add company name
    st.dataframe(st.session_state.balance_sheet_display_df, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.subheader("報表總計摘要")
    total_assets_val = totals.get('Assets', 0)
    total_liabilities_val = totals.get('Liabilities', 0)
    total_equity_val = totals.get('Equity', 0)
    total_liab_equity_val = total_liabilities_val + total_equity_val

    col1, col2 = st.columns(2)
    col1.metric("資產總計 (Total Assets)", f"{total_assets_val:,.0f}")
    col2.metric("負債及權益總計 (Total Liabilities and Equity)", f"{total_liab_equity_val:,.0f}")

    if round(total_assets_val) == round(total_liab_equity_val): # Use round for float comparisons
        st.success("會計方程式平衡：資產總計 = 負債及權益總計")
    else:
        st.error(f"會計方程式不平衡：資產總計 ({total_assets_val:,.0f}) != 負債及權益總計 ({total_liab_equity_val:,.0f})。差額: {total_assets_val - total_liab_equity_val:,.0f}")

    st.markdown("---")
    st.subheader("匯出功能")
    col_export_bs1, col_export_bs2 = st.columns(2)

    excel_data = bs_df_to_excel(st.session_state.balance_sheet_raw_df, params)
    col_export_bs1.download_button(
        label="📥 匯出 Excel", data=excel_data,
        file_name=f"BalanceSheet_{params.get('company_no','ALL')}_{params['as_of_date'].strftime('%Y%m%d')}.xlsx", # Add company
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    pdf_data = bs_df_to_pdf(st.session_state.balance_sheet_display_df, params)
    col_export_bs2.download_button(
        label="📄 匯出 PDF", data=pdf_data,
        file_name=f"BalanceSheet_{params.get('company_no','ALL')}_{params['as_of_date'].strftime('%Y%m%d')}.pdf", # Add company
        mime="application/pdf"
    )
else:
    st.info("請在側邊欄選擇公司和截止日期後，點擊「生成報表」。") # Updated prompt

st.sidebar.info("提示：此為簡化版資產負債表。詳細科目分類及保留盈餘的精確計算可能需要更複雜的科目設定和損益數據。")
