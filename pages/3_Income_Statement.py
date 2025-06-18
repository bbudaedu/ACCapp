import streamlit as st
import pandas as pd
from app.core.db_connector import get_db_engine
from sqlalchemy import text # Removed func
from calendar import monthrange
import datetime
import numpy as np # For abs and inf handling
import io # For BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF # For PDF export

# --- Page Configuration ---
st.set_page_config(page_title="損益表 (Income Statement)", layout="wide")
st.title("損益表 (Income Statement)")

# --- Database Connection ---
@st.cache_resource
def init_db_connection():
    engine = get_db_engine()
    return engine

db_engine = init_db_connection()
if not db_engine:
    st.error("資料庫引擎未能初始化。請檢查您的資料庫配置。")
    st.stop()

# --- Helper function to fetch data ---
@st.cache_data(ttl=3600) # Cache company data longer
def fetch_company_data_is():
    if not db_engine:
        return pd.DataFrame(columns=['CP_UNINO', 'CP_NAME']) # Use new column names
    try:
        with db_engine.connect() as connection:
            # Use new query: SELECT CP_UNINO, CP_NAME FROM PCOMPANY ORDER BY CP_NAME
            df = pd.read_sql(text("SELECT CP_UNINO, CP_NAME FROM PCOMPANY ORDER BY CP_NAME"), connection)
            return df
    except Exception as e:
        st.error(f"獲取公司列表時發生錯誤 (損益表): {e}")
        return pd.DataFrame(columns=['CP_UNINO', 'CP_NAME']) # Use new column names

@st.cache_data(ttl=300)
def execute_query(query, params=None):
    try:
        with db_engine.connect() as connection:
            df = pd.read_sql(text(query), connection, params=params)
            return df
    except Exception as e:
        st.error(f"資料查詢時發生錯誤: {e}")
        return pd.DataFrame()

# --- Income Statement Account Categories ---
IS_CATEGORIES_DEF = {
    "營業收入 (Revenue)": {"pattern": "4%", "type": "Revenue"},
    "營業成本 (COGS)": {"pattern": "5%", "type": "Expense"},
    "營業費用 (Operating Expenses)": {"pattern": "6%", "type": "Expense"},
    "營業外收入 (Non-operating Income)": {"pattern": "71%", "type": "Revenue"},
    "營業外支出 (Non-operating Expenses)": {"pattern": "75%", "type": "Expense"}
}
IS_DISPLAY_ORDER = [
    "營業收入 (Revenue)", "營業成本 (COGS)", "營業毛利 (Gross Profit)",
    "營業費用 (Operating Expenses)", "營業利益 (Operating Income)",
    "營業外收入 (Non-operating Income)", "營業外支出 (Non-operating Expenses)",
    "稅前淨利 (Pre-tax Income)"
]
CALCULATED_ITEMS = ["營業毛利 (Gross Profit)", "營業利益 (Operating Income)", "稅前淨利 (Pre-tax Income)"]

def fetch_is_category_data_for_period(year, month, company_no): # Added company_no
    if not company_no:
        st.sidebar.error("損益表錯誤：未選擇公司進行查詢。") # User feedback
        return {} # Return empty dict if no company
    _, num_days = monthrange(year, month)
    start_date = datetime.date(year, month, 1).strftime('%Y%m%d')
    end_date = datetime.date(year, month, num_days).strftime('%Y%m%d')
    period_data = {}
    for category_name, details in IS_CATEGORIES_DEF.items():
        acc_pattern = details["pattern"]
        query = """
        SELECT COALESCE(SUM(CASE WHEN d.SD_DOC = 'C' THEN d.SD_AMT ELSE -d.SD_AMT END), 0) AS Amount
        FROM ASPDT d
        JOIN ASLIP h ON d.SD_NO = h.SP_NO -- Removed d.SD_INDEX = h.SP_INDEX from join
        WHERE h.SP_CHECK = '1'
          AND d.SD_ATNO LIKE :acc_pattern
          AND h.SP_DATE BETWEEN :start_date AND :end_date
          AND h.SP_CO_NO = :company_no; -- Added company filter
        """
        params = {'acc_pattern': acc_pattern, 'start_date': start_date, 'end_date': end_date, 'company_no': company_no}
        df_amount = execute_query(query, params)
        amount = df_amount.iloc[0]['Amount'] if not df_amount.empty and 'Amount' in df_amount.columns else 0
        period_data[category_name] = amount
    return period_data

def calculate_derived_is_items(period_data_map): # No change to this function itself
    revenue = period_data_map.get("營業收入 (Revenue)", 0)
    cogs = period_data_map.get("營業成本 (COGS)", 0)
    period_data_map["營業毛利 (Gross Profit)"] = revenue + cogs
    op_expenses = period_data_map.get("營業費用 (Operating Expenses)", 0)
    period_data_map["營業利益 (Operating Income)"] = period_data_map["營業毛利 (Gross Profit)"] + op_expenses
    non_op_income = period_data_map.get("營業外收入 (Non-operating Income)", 0)
    non_op_expenses = period_data_map.get("營業外支出 (Non-operating Expenses)", 0)
    period_data_map["稅前淨利 (Pre-tax Income)"] = period_data_map["營業利益 (Operating Income)"] + non_op_income + non_op_expenses
    return period_data_map

def generate_comparative_is_df(current_period_data, current_period_label, ly_period_data=None, ly_period_label=None, lm_period_data=None, lm_period_label=None):
    df_rows = []
    for item_name in IS_DISPLAY_ORDER:
        row = {"項目 (Item)": item_name}
        current_amount = current_period_data.get(item_name, 0)

        # For display, expenses are positive. For calculation, their original sign is used.
        is_expense_item = item_name in IS_CATEGORIES_DEF and IS_CATEGORIES_DEF[item_name]["type"] == "Expense"

        row[current_period_label] = -current_amount if is_expense_item else current_amount

        if ly_period_data and ly_period_label:
            ly_amount = ly_period_data.get(item_name, 0)
            row[ly_period_label] = -ly_amount if is_expense_item else ly_amount
            if ly_amount != 0:
                change = (current_amount - ly_amount) / abs(ly_amount) * 100
                row["年增率 YoY (%)"] = change
            elif current_amount != 0:
                row["年增率 YoY (%)"] = np.inf # Indicates new item or from zero
            else:
                row["年增率 YoY (%)"] = 0 # Both zero

        if lm_period_data and lm_period_label:
            lm_amount = lm_period_data.get(item_name, 0)
            row[lm_period_label] = -lm_amount if is_expense_item else lm_amount
            if lm_amount != 0:
                change = (current_amount - lm_amount) / abs(lm_amount) * 100
                row["月增率 MoM (%)"] = change
            elif current_amount != 0:
                row["月增率 MoM (%)"] = np.inf
            else:
                row["月增率 MoM (%)"] = 0
        df_rows.append(row)
    return pd.DataFrame(df_rows)


def df_to_excel_with_formulas(df, params):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"IncomeStatement_{params['year']}{params['month']:02d}"

    # Add title
    ws.append([f"{params['year']} 年 {params['month']:02d} 月 損益表"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1])
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([]) # Empty row

    # Write DataFrame header
    header = [cell for cell in df.columns]
    ws.append(header)
    for cell in ws[3:3]: # Header row is 3
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Write data rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 4): # Data starts at row 4
        ws.append(row)
        # Apply formulas for calculated items
        item_name = row[0] # Assuming '項目 (Item)' is the first column
        # This is tricky because df has formatted strings for %
        # We need to find column indices for amounts
        amount_cols_indices = {col_name: idx + 1 for idx, col_name in enumerate(df.columns) if "金額" in col_name}

        if item_name == "營業毛利 (Gross Profit)":
            for col_letter_idx_str, amount_col_idx in amount_cols_indices.items():
                col_letter = get_column_letter(amount_col_idx)
                revenue_row_idx = df[df['項目 (Item)'] == "營業收入 (Revenue)"].index[0] + 4 # +4 for title, space, header
                cogs_row_idx = df[df['項目 (Item)'] == "營業成本 (COGS)"].index[0] + 4
                # Formula: Revenue - COGS (since COGS is displayed positive but represents expense)
                ws[f'{col_letter}{r_idx}'] = f"={col_letter}{revenue_row_idx}-{col_letter}{cogs_row_idx}"
        elif item_name == "營業利益 (Operating Income)":
            for col_letter_idx_str, amount_col_idx in amount_cols_indices.items():
                col_letter = get_column_letter(amount_col_idx)
                gp_row_idx = df[df['項目 (Item)'] == "營業毛利 (Gross Profit)"].index[0] + 4
                opex_row_idx = df[df['項目 (Item)'] == "營業費用 (Operating Expenses)"].index[0] + 4
                ws[f'{col_letter}{r_idx}'] = f"={col_letter}{gp_row_idx}-{col_letter}{opex_row_idx}"
        elif item_name == "稅前淨利 (Pre-tax Income)":
            for col_letter_idx_str, amount_col_idx in amount_cols_indices.items():
                col_letter = get_column_letter(amount_col_idx)
                op_income_row_idx = df[df['項目 (Item)'] == "營業利益 (Operating Income)"].index[0] + 4
                non_op_income_row_idx = df[df['項目 (Item)'] == "營業外收入 (Non-operating Income)"].index[0] + 4
                non_op_expense_row_idx = df[df['項目 (Item)'] == "營業外支出 (Non-operating Expenses)"].index[0] + 4
                ws[f'{col_letter}{r_idx}'] = f"={col_letter}{op_income_row_idx}+{col_letter}{non_op_income_row_idx}-{col_letter}{non_op_expense_row_idx}"

        # Style calculated rows
        if item_name in CALCULATED_ITEMS:
            for cell in ws[r_idx:r_idx]:
                cell.font = Font(bold=True)

    # Number formatting and column widths (example)
    for col_idx, column_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        if "金額" in column_name:
            for r in range(4, ws.max_row + 1): # Data rows
                ws[f'{col_letter}{r}'].number_format = '#,##0'
            ws.column_dimensions[col_letter].width = 18
        elif "%" in column_name:
             for r in range(4, ws.max_row + 1):
                # Need to handle 'N/A', '新產生' before trying to format as percentage
                if isinstance(ws[f'{col_letter}{r}'].value, (int, float)) and ws[f'{col_letter}{r}'].value != np.inf :
                     ws[f'{col_letter}{r}'].number_format = '0.0"%"'
                     ws[f'{col_letter}{r}'].value = ws[f'{col_letter}{r}'].value / 100 # Store as actual number for Excel %
                elif ws[f'{col_letter}{r}'].value == np.inf:
                     ws[f'{col_letter}{r}'].value = "新產生"

            ws.column_dimensions[col_letter].width = 15
        else: # Item column
            ws.column_dimensions[col_letter].width = 30

    wb.save(output)
    output.seek(0)
    return output.getvalue()

def df_to_pdf(df, params):
    pdf = FPDF()
    pdf.add_page()

    # Add font that supports CJK characters (assuming the environment has one)
    # This is a common issue. If no CJK font is found, it will use a default one that might not render Chinese.
    # For this example, let's assume a font 'uming.ttc' or 'msjh.ttc' might be available in a sandbox.
    # This path is a placeholder and likely needs adjustment for the execution environment.
    try:
        pdf.add_font('uming', '', '/usr/share/fonts/truetype/arphic/uming.ttc', uni=True) # Common path in Linux
        pdf.set_font('uming', '', 10)
    except RuntimeError:
        try:
            pdf.add_font('msjh', '', 'C:/Windows/Fonts/msjh.ttc', uni=True) # Common path in Windows
            pdf.set_font('msjh', '', 10)
        except RuntimeError:
            st.warning("CJK font not found for PDF export. Chinese characters might not render correctly.")
            pdf.set_font('Arial', '', 10)


    pdf.cell(0, 10, f"{params['year']} 年 {params['month']:02d} 月 損益表", 0, 1, 'C')

    # Table Header
    header = list(df.columns)
    col_widths = {'項目 (Item)': 70} # Estimate widths
    for h in header:
        if h not in col_widths: col_widths[h] = 30

    line_height = pdf.font_size * 1.5

    for idx, h_text in enumerate(header):
        pdf.cell(col_widths.get(h_text, 30), line_height, h_text, border=1, ln=0 if idx < len(header)-1 else 1, align='C')

    # Table Rows
    for _, row_series in df.iterrows():
        for idx, (col_name, cell_value) in enumerate(row_series.items()):
            if isinstance(cell_value, (float, int)) and col_name not in ["年增率 YoY (%)", "月增率 MoM (%)"]:
                cell_text = f"{cell_value:,.0f}"
            elif isinstance(cell_value, float) and col_name in ["年增率 YoY (%)", "月增率 MoM (%)"]:
                if cell_value == np.inf: cell_text = "新產生"
                else: cell_text = f"{cell_value:.1f}%"
            else:
                cell_text = str(cell_value)
            pdf.cell(col_widths.get(col_name, 30), line_height, cell_text, border=1, ln=0 if idx < len(header)-1 else 1, align='R' if isinstance(cell_value, (float, int)) else 'L')

    return pdf.output(dest='S').encode('latin-1') # S returns bytes


# --- Filter Section ---
st.sidebar.header("報表參數")

# Company Selector for Income Statement
companies_df_is = fetch_company_data_is()
# Update to use CP_UNINO and CP_NAME
company_options_is = {row['CP_UNINO']: row['CP_NAME'] for _, row in companies_df_is.iterrows()} if not companies_df_is.empty else {}
selected_company_unino_is = None # Changed variable name
selected_company_name_is = "無公司"

if not company_options_is:
    st.sidebar.warning("未找到任何公司資料 (損益表)。")
else:
    default_company_unino_is = list(company_options_is.keys())[0]
    selected_company_unino_is = st.sidebar.selectbox(
        "公司 (Company)", options=list(company_options_is.keys()), # Options are CP_UNINO
        format_func=lambda x: company_options_is.get(x, "未知公司"), # Format uses CP_NAME
        key="is_company_unino", index=0 # Changed key
    )
    selected_company_name_is = company_options_is.get(selected_company_unino_is, "未知公司")


current_year_today = datetime.date.today().year
current_month_today = datetime.date.today().month
years_list = list(range(current_year_today - 10, current_year_today + 1))
selected_year = st.sidebar.selectbox("選擇年份", options=years_list, index=len(years_list)-1, key="is_year")
months_list = list(range(1, 13))
month_names_map = {m: datetime.date(2000, m, 1).strftime('%B') for m in months_list}
selected_month = st.sidebar.selectbox("選擇月份", options=months_list, format_func=lambda m: f"{m:02d} ({month_names_map[m]})", index=current_month_today-1, key="is_month")
st.sidebar.subheader("比較功能")
compare_ly_cb = st.sidebar.checkbox("與去年同期比較 (%)", key="compare_ly_is")
compare_lm_cb = st.sidebar.checkbox("與上月比較 (%)", key="compare_lm_is")

# --- Session State ---
if 'income_statement_final_df' not in st.session_state:
    st.session_state.income_statement_final_df = pd.DataFrame()
if 'income_statement_raw_data_for_export' not in st.session_state:
    st.session_state.income_statement_raw_data_for_export = pd.DataFrame()
if 'income_statement_display_params' not in st.session_state:
    st.session_state.income_statement_display_params = None

# --- Generate Report Button ---
if st.sidebar.button("生成報表", type="primary", key="is_generate_button"):
    if not selected_company_unino_is: # Check updated variable
        st.error("請選擇一個公司。")
    else:
        st.session_state.income_statement_display_params = {
            "company_name": selected_company_name_is, "company_unino": selected_company_unino_is, # Store CP_UNINO
            "year": selected_year, "month": selected_month,
            "compare_ly": compare_ly_cb, "compare_lm": compare_lm_cb
        }
        with st.spinner(f"正在為 {selected_company_name_is} 生成 {selected_year} 年 {selected_month:02d} 月損益表..."):
            current_data_raw_cats = fetch_is_category_data_for_period(selected_year, selected_month, selected_company_unino_is) # Pass CP_UNINO
            current_data_calculated = calculate_derived_is_items(current_data_raw_cats.copy())

            ly_data_calculated, lm_data_calculated = None, None
            ly_label_for_df, lm_label_for_df = None, None
            ly_year, ly_month = selected_year - 1, selected_month
            lm_year, lm_month = (selected_year, selected_month - 1) if selected_month > 1 else (selected_year - 1, 12)

            if compare_ly_cb:
                ly_label_for_df = f"金額 ({ly_year}/{ly_month:02d} LY)"
                ly_data_raw_cats = fetch_is_category_data_for_period(ly_year, ly_month, selected_company_unino_is) # Pass CP_UNINO
                ly_data_calculated = calculate_derived_is_items(ly_data_raw_cats.copy())
            if compare_lm_cb:
                lm_label_for_df = f"金額 ({lm_year}/{lm_month:02d} LM)"
                lm_data_raw_cats = fetch_is_category_data_for_period(lm_year, lm_month, selected_company_unino_is) # Pass CP_UNINO
                lm_data_calculated = calculate_derived_is_items(lm_data_raw_cats.copy())

        current_period_label_for_df = f"{selected_year}/{selected_month:02d} 金額"

        # This df is for export, contains raw numbers for % change columns
        df_for_export = generate_comparative_is_df(
            current_data_calculated, current_period_label_for_df,
            ly_data_calculated, ly_label_for_df,
            lm_data_calculated, lm_label_for_df
        )
        st.session_state.income_statement_raw_data_for_export = df_for_export.copy() # Save for export

        # Create a separate DataFrame for display with formatted strings
        df_for_display = df_for_export.copy()
        amount_cols = [col for col in df_for_display.columns if "金額" in col]
        for col in amount_cols:
            df_for_display[col] = df_for_display[col].apply(lambda x: f"{x:,.0f}" if pd.notnull(x) else "0")

        percent_cols = [col for col in df_for_display.columns if "%" in col]
        for col in percent_cols:
            df_for_display[col] = df_for_display[col].apply(
                lambda x: f"{x:.1f}%" if pd.notnull(x) and x != np.inf else ("新產生" if x == np.inf else "N/A")
            )
        st.session_state.income_statement_final_df = df_for_display


# --- Display Area for Income Statement ---
if not st.session_state.income_statement_final_df.empty and st.session_state.income_statement_display_params:
    params = st.session_state.income_statement_display_params
    # Display company name in the report title
    st.subheader(f"{params.get('company_name', '')} - {params['year']} 年 {params['month']:02d} 月 損益表")
    st.dataframe(st.session_state.income_statement_final_df, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.subheader("匯出功能")
    col_export1, col_export2 = st.columns(2)

    excel_data = df_to_excel_with_formulas(st.session_state.income_statement_raw_data_for_export, params)
    col_export1.download_button(
        label="📥 匯出 Excel", data=excel_data,
        file_name=f"IncomeStatement_{params.get('company_unino','ALL')}_{params['year']}{params['month']:02d}.xlsx", # Use company_unino
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    pdf_data = df_to_pdf(st.session_state.income_statement_final_df, params)
    col_export2.download_button(
        label="📄 匯出 PDF", data=pdf_data,
        file_name=f"IncomeStatement_{params.get('company_unino','ALL')}_{params['year']}{params['month']:02d}.pdf", # Use company_unino
        mime="application/pdf"
    )
else:
    st.info("請在側邊欄選擇公司、年份和月份後，點擊「生成報表」。")

st.sidebar.info("注意：會計科目範圍需與客戶確認。PDF匯出需要環境中包含支援中文的字型。")
